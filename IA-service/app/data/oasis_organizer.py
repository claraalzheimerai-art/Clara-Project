"""
Organiza los datos de OASIS-1 en la estructura esperada por el pipeline.
Compatible con formato Analyze (.img/.hdr) y Excel (.xlsx).

Uso:
    python -m app.data.oasis_organizer \
        --xlsx_path "C:\\Users\\Nahia Sanchez\\Desktop\\Data-clara\\oasis_cross-sectional-5708aa0a98d82080.xlsx" \
        --disc_dirs "C:\\Users\\Nahia Sanchez\\Desktop\\Data-clara\\disc1" "C:\\Users\\Nahia Sanchez\\Desktop\\Data-clara\\disc2" \
        --output_dir app/data/raw
"""
import argparse
import shutil
from pathlib import Path
from app.core.logger import logger

try:
    import openpyxl
except ImportError:
    raise ImportError("Ejecuta: pip install openpyxl")


CDR_TO_LABEL = {
    0:    "CN",
    0.5:  "MCI",
    1.0:  "AD",
    2.0:  "AD",
}


def read_xlsx_labels(xlsx_path: str) -> dict:
    """
    Lee el Excel de OASIS-1 y retorna {subject_id: label}.
    Toma la primera fila con CDR válido por sujeto.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" 
               for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    try:
        id_col  = headers.index("ID")
        cdr_col = headers.index("CDR")
    except ValueError:
        raise ValueError(f"Columnas esperadas 'ID' y 'CDR' no encontradas. Encontradas: {headers}")

    subject_labels = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_id = row[id_col]
        cdr    = row[cdr_col]

        if not raw_id:
            continue

        subject_id = _normalize_id(str(raw_id).strip())

        if cdr is not None and float(cdr) in CDR_TO_LABEL:
            if subject_id not in subject_labels:
                subject_labels[subject_id] = CDR_TO_LABEL[float(cdr)]

    wb.close()
    logger.info(f"Sujetos con CDR válido: {len(subject_labels)}")

    from collections import Counter
    counts = Counter(subject_labels.values())
    logger.info(f"Distribución en CSV: CN={counts['CN']} | MCI={counts['MCI']} | AD={counts['AD']}")

    return subject_labels


def organize_oasis1(xlsx_path: str, disc_dirs: list, output_dir: str):
    out_root = Path(output_dir)
    for label in ["CN", "MCI", "AD"]:
        (out_root / label).mkdir(parents=True, exist_ok=True)

    # Leer etiquetas del Excel
    subject_labels = read_xlsx_labels(xlsx_path)

    # Indexar todos los archivos .img disponibles en los discos
    logger.info("Indexando archivos .img en los discos...")
    img_index = {}
    for disc_dir in disc_dirs:
        disc_path = Path(disc_dir)
        for img_file in disc_path.rglob("*T88_111/*.img"):
            # Extraer subject_id del nombre del archivo
            # OAS1_0001_MR1_mpr_n4_anon_111_t88_gfc.img → OAS1_0001
            parts = img_file.name.split("_")
            if len(parts) >= 2:
                sid = f"{parts[0]}_{parts[1]}"
                # Preferir masked sobre no-masked
                if sid not in img_index or "masked" in img_file.name:
                    img_index[sid] = img_file

    logger.info(f"Archivos .img indexados: {len(img_index)}")

    # Copiar archivos a carpetas por clase
    stats = {"CN": 0, "MCI": 0, "AD": 0, "no_encontrado": 0}

    for subject_id, label in subject_labels.items():
        img_file = img_index.get(subject_id)

        if img_file is None:
            logger.warning(f"No encontrado: {subject_id}")
            stats["no_encontrado"] += 1
            continue

        # Copiar .img y su .hdr correspondiente
        hdr_file = img_file.with_suffix(".hdr")
        dest_img = out_root / label / f"{subject_id}.img"
        dest_hdr = out_root / label / f"{subject_id}.hdr"

        if not dest_img.exists():
            shutil.copy2(str(img_file), str(dest_img))
        if hdr_file.exists() and not dest_hdr.exists():
            shutil.copy2(str(hdr_file), str(dest_hdr))

        stats[label] += 1
        logger.info(f"{subject_id} → {label}")

    logger.info("=" * 40)
    logger.info(f"Resultado: {stats}")
    total = stats["CN"] + stats["MCI"] + stats["AD"]
    logger.info(f"Total organizado: {total} sujetos")
    return stats


def _normalize_id(raw_id: str) -> str:
    parts = raw_id.split("_")
    if len(parts) >= 2:
        return f"{parts[0]}_{parts[1]}"
    return raw_id


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx_path",  required=True)
    parser.add_argument("--disc_dirs",  nargs="+", required=True)
    parser.add_argument("--output_dir", default="app/data/raw")
    args = parser.parse_args()

    organize_oasis1(args.xlsx_path, args.disc_dirs, args.output_dir)